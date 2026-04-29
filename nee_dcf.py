"""
NextEra Energy (NEE) – Discounted Cash Flow Model
Methodology note: NEE is a capital-intensive regulated utility whose growth capex
($9B+/yr) creates regulated-return assets that yfinance's FCF metric understates.
We use UFCF = EBIT × (1 − effective tax rate), which normalises for growth capex
and is consistent with the revenue-growth projections (new assets generate new revenue).
"""

import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import warnings
warnings.filterwarnings("ignore")

TICKER = "NEE"
PROJECTION_YEARS = 5
RISK_FREE_RATE = 0.043
EQUITY_RISK_PREMIUM = 0.055
TERMINAL_GROWTH_RATE = 0.030      # 3% – utility regulated growth
REVENUE_GROWTH_RATES = [0.08, 0.08, 0.07, 0.07, 0.06]

# ── palette ───────────────────────────────────────────────────────────────────
DARK_GREEN  = "1F4E35"
MID_GREEN   = "2E7D52"
LIGHT_GREEN = "D9EAD3"
BORDER_COL  = "CCCCCC"

def _side(): return Side(style="thin", color=BORDER_COL)
def all_border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
def hfill(col=DARK_GREEN): return PatternFill("solid", fgColor=col)
def lfill(): return PatternFill("solid", fgColor=LIGHT_GREEN)
def gfill(): return PatternFill("solid", fgColor="F5F5F5")

# ── fetch ─────────────────────────────────────────────────────────────────────
def fetch_financials(ticker):
    tk   = yf.Ticker(ticker)
    info = tk.info

    inc = tk.financials
    cf  = tk.cashflow
    bs  = tk.balance_sheet

    inc = inc[sorted(inc.columns)].iloc[:, -3:]
    cf  = cf[sorted(cf.columns)].iloc[:, -3:]
    bs  = bs[sorted(bs.columns)]

    years = [str(c.year) for c in inc.columns]

    def get(df, *keys):
        for k in keys:
            for idx in df.index:
                if k.lower() in idx.lower():
                    return df.loc[idx].values.astype(float)
        return np.zeros(len(inc.columns))

    def bget(df, *keys):
        for k in keys:
            for idx in df.index:
                if k.lower() in idx.lower():
                    return float(df.loc[idx].iloc[-1])
        return 0.0

    revenue    = get(inc, "Total Revenue")
    ebit       = get(inc, "EBIT")
    ebitda     = get(inc, "Normalized EBITDA", "EBITDA")
    da         = get(inc, "Reconciled Depreciation", "Depreciation And Amortization")
    net_income = get(inc, "Net Income Common")
    pretax     = get(inc, "Pretax Income")
    tax        = get(inc, "Tax Provision")
    interest_e = get(inc, "Interest Expense Non Operating", "Interest Expense")
    op_cf      = get(cf,  "Operating Cash Flow")
    capex_raw  = get(cf,  "Capital Expenditure")
    raw_fcf    = get(cf,  "Free Cash Flow")

    with np.errstate(divide="ignore", invalid="ignore"):
        eff_tax = np.where(pretax != 0, np.abs(tax) / np.abs(pretax), 0.10)
    eff_tax = np.clip(eff_tax, 0.05, 0.30)
    avg_tax = float(np.mean(eff_tax[-2:]))

    # UFCF = EBIT × (1 − t)  [growth-capex-normalised for regulated utility]
    nopat = ebit * (1 - avg_tax)

    total_debt    = bget(bs, "Total Debt", "Long Term Debt")
    cash          = bget(bs, "Cash And Cash Equivalents")
    shares_out    = float(info.get("sharesOutstanding") or info.get("impliedSharesOutstanding", 2e9))
    current_price = float(info.get("currentPrice") or info.get("regularMarketPrice", 100))
    market_cap    = current_price * shares_out
    beta          = float(info.get("beta") or 0.65)

    avg_interest_e = float(np.mean(np.abs(interest_e[-2:])))

    return dict(
        years=years, revenue=revenue, ebit=ebit, ebitda=ebitda,
        da=da, net_income=net_income, nopat=nopat,
        op_cf=op_cf, capex=capex_raw, raw_fcf=raw_fcf,
        avg_tax=avg_tax, eff_tax=eff_tax,
        total_debt=total_debt, cash=cash,
        shares_out=shares_out, current_price=current_price,
        market_cap=market_cap, beta=beta,
        avg_interest_e=avg_interest_e,
    )

# ── projections ───────────────────────────────────────────────────────────────
def build_projections(hist):
    last_rev   = hist["revenue"][-1]
    last_ebit  = hist["ebit"][-1]
    ebit_margin_base = last_ebit / last_rev
    # NEE EBIT margin has been 30-35%; project with slight expansion
    target_margin = 0.34
    margin_step   = (target_margin - ebit_margin_base) / PROJECTION_YEARS

    proj_years, proj_rev, proj_ebit, proj_nopat, proj_margins = [], [], [], [], []
    rev = last_rev
    for i, g in enumerate(REVENUE_GROWTH_RATES):
        rev   = rev * (1 + g)
        marg  = ebit_margin_base + margin_step * (i + 1)
        ebit  = rev * marg
        nopat = ebit * (1 - hist["avg_tax"])
        proj_years.append(str(int(hist["years"][-1]) + i + 1))
        proj_rev.append(rev)
        proj_ebit.append(ebit)
        proj_nopat.append(nopat)
        proj_margins.append(marg)

    return dict(
        proj_years=proj_years, proj_rev=proj_rev, proj_ebit=proj_ebit,
        proj_nopat=proj_nopat, proj_margins=proj_margins,
        growth_rates=REVENUE_GROWTH_RATES,
    )

# ── WACC ──────────────────────────────────────────────────────────────────────
def calc_wacc(hist):
    mc   = hist["market_cap"]
    debt = hist["total_debt"]
    beta = hist["beta"]
    tx   = hist["avg_tax"]
    total_capital = mc + debt

    we  = mc / total_capital
    wd  = debt / total_capital
    ke  = RISK_FREE_RATE + beta * EQUITY_RISK_PREMIUM
    kd  = np.clip(hist["avg_interest_e"] / debt if debt else 0.04, 0.02, 0.08)
    wac = we * ke + wd * kd * (1 - tx)

    return dict(wacc=wac, ke=ke, kd=kd, we=we, wd=wd,
                market_cap=mc, total_debt=debt, beta=beta, avg_tax=tx)

# ── DCF ───────────────────────────────────────────────────────────────────────
def calc_dcf(proj, wacc_data, hist):
    w    = wacc_data["wacc"]
    fcfs = proj["proj_nopat"]

    pv_fcfs = [f / (1 + w) ** (i + 1) for i, f in enumerate(fcfs)]
    sum_pv  = sum(pv_fcfs)

    tv     = fcfs[-1] * (1 + TERMINAL_GROWTH_RATE) / (w - TERMINAL_GROWTH_RATE)
    pv_tv  = tv / (1 + w) ** PROJECTION_YEARS

    ev     = sum_pv + pv_tv
    equity = ev - hist["total_debt"] + hist["cash"]
    price  = equity / hist["shares_out"]

    return dict(pv_fcfs=pv_fcfs, sum_pv=sum_pv, terminal_value=tv,
                pv_terminal=pv_tv, enterprise_value=ev,
                equity_value=equity, implied_price=price,
                terminal_growth=TERMINAL_GROWTH_RATE)

# ── sensitivity ───────────────────────────────────────────────────────────────
def sensitivity_table(proj, wacc_base, hist):
    wacc_range = [wacc_base + d for d in (-0.02, -0.01, 0, 0.01, 0.02)]
    tgr_range  = [0.010, 0.015, 0.020, 0.025, 0.030, 0.035, 0.040]
    fcfs = proj["proj_nopat"]
    table = {}
    for tg in tgr_range:
        row = {}
        for wc in wacc_range:
            if wc <= tg:
                row[wc] = float("nan")
                continue
            pv   = sum(f / (1 + wc) ** (i + 1) for i, f in enumerate(fcfs))
            tv   = fcfs[-1] * (1 + tg) / (wc - tg)
            ptv  = tv / (1 + wc) ** PROJECTION_YEARS
            ev   = pv + ptv
            eq   = ev - hist["total_debt"] + hist["cash"]
            row[wc] = eq / hist["shares_out"]
        table[tg] = row
    return table, wacc_range, tgr_range

# ── Excel ─────────────────────────────────────────────────────────────────────
def _hdr(ws, r, c, v, bg=DARK_GREEN, fg="FFFFFF", bold=True, sz=10):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(bold=bold, color=fg, size=sz, name="Calibri")
    cell.fill      = hfill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = all_border()
    return cell

def _cell(ws, r, c, v, fmt=None, bold=False, fill=None, align="right"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(bold=bold, size=10, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = all_border()
    if fmt:  cell.number_format = fmt
    if fill: cell.fill = fill
    return cell

def _lbl(ws, r, c, v, bold=False, fill=None):
    return _cell(ws, r, c, v, bold=bold, fill=fill, align="left")

def write_excel(hist, proj, wacc_data, dcf, sens_table, sens_wacc, sens_tgr, path):
    wb = Workbook()

    # ═══════════════════════════════ SHEET 1 ════════════════════════════════
    ws = wb.active
    ws.title = "DCF Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "NextEra Energy (NEE) — Discounted Cash Flow Model"
    t.font  = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    t.fill  = hfill(DARK_GREEN)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value = (f"As of {pd.Timestamp.today().strftime('%B %d, %Y')}  |  "
               f"WACC: {wacc_data['wacc']:.2%}  |  Terminal Growth: {TERMINAL_GROWTH_RATE:.2%}  |  "
               f"FCF = NOPAT (EBIT × [1−t])  — Utility-Normalised")
    s.font  = Font(size=9, color="FFFFFF", name="Calibri")
    s.fill  = hfill(MID_GREEN)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Historical ──────────────────────────────────────────────────────────
    r = 4
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value="HISTORICAL FINANCIALS  ($ billions)")
    c.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    c.fill = hfill(MID_GREEN)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18

    r += 1
    _hdr(ws, r, 1, "Metric")
    for i, yr in enumerate(hist["years"]):
        _hdr(ws, r, i + 2, yr)

    hist_rows = [
        ("Revenue",              hist["revenue"],    "#,##0.0",  False),
        ("EBITDA",               hist["ebitda"],     "#,##0.0",  False),
        ("D&A",                  hist["da"],         "#,##0.0",  False),
        ("EBIT",                 hist["ebit"],       "#,##0.0",  False),
        ("NOPAT  (EBIT×[1−t])",  hist["nopat"],     "#,##0.0",  True),
        ("Operating Cash Flow",  hist["op_cf"],      "#,##0.0",  False),
        ("Capital Expenditure",  hist["capex"],      "#,##0.0",  False),
        ("Reported Free CF",     hist["raw_fcf"],    "#,##0.0",  False),
        ("Net Income",           hist["net_income"], "#,##0.0",  False),
    ]
    for lbl, vals, fmt, bold in hist_rows:
        r += 1
        fl = lfill() if bold else (gfill() if r % 2 == 0 else None)
        _lbl(ws, r, 1, lbl, bold=bold, fill=fl)
        for i, v in enumerate(vals):
            _cell(ws, r, i + 2, v / 1e9, fmt=fmt, bold=bold, fill=fl)

    # ── Projections ──────────────────────────────────────────────────────────
    r += 2
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value="5-YEAR PROJECTIONS  ($ billions)")
    c.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    c.fill = hfill(MID_GREEN)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18

    r += 1
    _hdr(ws, r, 1, "Metric")
    for i, yr in enumerate(proj["proj_years"]):
        _hdr(ws, r, i + 2, yr)

    proj_rows = [
        ("Revenue Growth",          proj["growth_rates"],                        "0.0%",    False),
        ("Revenue",                 [v / 1e9 for v in proj["proj_rev"]],        "#,##0.0", False),
        ("EBIT Margin",             proj["proj_margins"],                        "0.0%",    False),
        ("EBIT",                    [v / 1e9 for v in proj["proj_ebit"]],       "#,##0.0", False),
        ("NOPAT  (FCF proxy)",      [v / 1e9 for v in proj["proj_nopat"]],      "#,##0.0", True),
        ("PV of NOPAT",             [v / 1e9 for v in dcf["pv_fcfs"]],         "#,##0.0", True),
    ]
    for lbl, vals, fmt, bold in proj_rows:
        r += 1
        fl = lfill() if bold else (gfill() if r % 2 == 0 else None)
        _lbl(ws, r, 1, lbl, bold=bold, fill=fl)
        for i, v in enumerate(vals):
            _cell(ws, r, i + 2, v, fmt=fmt, bold=bold, fill=fl)

    # ── Valuation Bridge ──────────────────────────────────────────────────────
    r += 2
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value="VALUATION BRIDGE  ($ billions unless noted)")
    c.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    c.fill = hfill(MID_GREEN)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18

    bridge = [
        ("Sum of PV (Years 1-5)",           dcf["sum_pv"] / 1e9,              "#,##0.0",      False),
        ("Terminal Value (Gordon Growth)",   dcf["terminal_value"] / 1e9,      "#,##0.0",      False),
        ("PV of Terminal Value",             dcf["pv_terminal"] / 1e9,         "#,##0.0",      True),
        ("Enterprise Value",                 dcf["enterprise_value"] / 1e9,    "#,##0.0",      True),
        ("Less: Total Debt",                -hist["total_debt"] / 1e9,         "(#,##0.0)",     False),
        ("Plus: Cash & Equivalents",         hist["cash"] / 1e9,               "#,##0.0",      False),
        ("Equity Value",                     dcf["equity_value"] / 1e9,        "#,##0.0",      True),
        ("Shares Outstanding (millions)",    hist["shares_out"] / 1e6,         "#,##0.0",      False),
        ("Implied Share Price",              dcf["implied_price"],              '"$"#,##0.00',  True),
        ("Current Share Price",              hist["current_price"],             '"$"#,##0.00',  False),
        ("Premium / (Discount) to Current",
         (dcf["implied_price"] - hist["current_price"]) / hist["current_price"],
         '+0.0%;-0.0%;0.0%', True),
    ]
    for lbl, val, fmt, bold in bridge:
        r += 1
        fl = lfill() if bold else (gfill() if r % 2 == 0 else None)
        _lbl(ws, r, 1, lbl, bold=bold, fill=fl)
        _cell(ws, r, 2, val, fmt=fmt, bold=bold, fill=fl)

    # ── WACC ──────────────────────────────────────────────────────────────────
    r += 2
    ws.merge_cells(f"A{r}:H{r}")
    c = ws.cell(row=r, column=1, value="WACC CALCULATION")
    c.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    c.fill = hfill(MID_GREEN)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18

    wacc_rows = [
        ("Risk-Free Rate (10-yr UST)",   RISK_FREE_RATE,             "0.00%",  False),
        ("Equity Risk Premium",          EQUITY_RISK_PREMIUM,        "0.00%",  False),
        ("Beta",                         wacc_data["beta"],           "0.000",  False),
        ("Cost of Equity  (CAPM)",       wacc_data["ke"],            "0.00%",  False),
        ("Pre-Tax Cost of Debt",         wacc_data["kd"],            "0.00%",  False),
        ("Effective Tax Rate",           wacc_data["avg_tax"],       "0.00%",  False),
        ("Weight of Equity",             wacc_data["we"],            "0.00%",  False),
        ("Weight of Debt",               wacc_data["wd"],            "0.00%",  False),
        ("WACC",                         wacc_data["wacc"],          "0.00%",  True),
    ]
    for lbl, val, fmt, bold in wacc_rows:
        r += 1
        fl = lfill() if bold else (gfill() if r % 2 == 0 else None)
        _lbl(ws, r, 1, lbl, bold=bold, fill=fl)
        _cell(ws, r, 2, val, fmt=fmt, bold=bold, fill=fl)

    # col widths
    ws.column_dimensions["A"].width = 38
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 17

    # ═══════════════════════════════ SHEET 2 ════════════════════════════════
    ws2 = wb.create_sheet("Sensitivity Analysis")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value = "NEE DCF — Sensitivity: Implied Share Price  (WACC vs Terminal Growth Rate)"
    t2.font  = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    t2.fill  = hfill(DARK_GREEN)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:G2")
    s2 = ws2["A2"]
    s2.value = (f"Base WACC: {wacc_data['wacc']:.2%}  |  "
                f"Base TGR: {TERMINAL_GROWTH_RATE:.2%}  |  "
                f"Current Price: ${hist['current_price']:.2f}  |  "
                "Shaded cell = base case")
    s2.font  = Font(size=9, color="FFFFFF", name="Calibri")
    s2.fill  = hfill(MID_GREEN)
    s2.alignment = Alignment(horizontal="center", vertical="center")

    r2 = 4
    corner = ws2.cell(row=r2, column=1, value="TGR  ↓   WACC  →")
    corner.font      = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
    corner.fill      = hfill(DARK_GREEN)
    corner.alignment = Alignment(horizontal="center", vertical="center")
    corner.border    = all_border()

    for j, wc in enumerate(sens_wacc):
        c = ws2.cell(row=r2, column=j + 2, value=wc)
        c.font         = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        c.fill         = hfill(DARK_GREEN)
        c.number_format = "0.00%"
        c.alignment    = Alignment(horizontal="center", vertical="center")
        c.border       = all_border()

    for i, tg in enumerate(sens_tgr):
        r2 += 1
        is_base_row = abs(tg - TERMINAL_GROWTH_RATE) < 1e-9
        for j, wc in enumerate(sens_wacc):
            is_base_cell = is_base_row and abs(wc - wacc_data["wacc"]) < 1e-9
            price = sens_table[tg][wc]

            if j == 0:  # row label
                lc = ws2.cell(row=r2, column=1, value=tg)
                lc.number_format = "0.0%"
                lc.font          = Font(bold=is_base_row, size=10, name="Calibri")
                lc.fill          = lfill() if is_base_row else (gfill() if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF"))
                lc.alignment     = Alignment(horizontal="center", vertical="center")
                lc.border        = all_border()

            val_c = ws2.cell(row=r2, column=j + 2,
                             value=price if not np.isnan(price) else "N/A")
            if not np.isnan(price):
                val_c.number_format = '"$"#,##0.00'
            val_c.font      = Font(bold=is_base_cell, size=10, name="Calibri",
                                   color=DARK_GREEN if is_base_cell else "000000")
            val_c.fill      = (lfill() if is_base_cell
                               else (gfill() if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")))
            val_c.alignment = Alignment(horizontal="center", vertical="center")
            val_c.border    = all_border()

    price_range = f"B5:F{r2}"
    ws2.conditional_formatting.add(
        price_range,
        ColorScaleRule(
            start_type="min",        start_color="FF6B6B",
            mid_type="percentile",   mid_value=50, mid_color="FFEB84",
            end_type="max",          end_color="63BE7B",
        )
    )

    ws2.column_dimensions["A"].width = 12
    for col in range(2, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    # ═══════════════════════════════ SHEET 3 ════════════════════════════════
    ws3 = wb.create_sheet("Assumptions")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value = "Model Assumptions & Methodology Notes"
    t3.font  = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    t3.fill  = hfill(DARK_GREEN)
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    notes = [
        ("Parameter",              "Value",                              "Note"),
        ("FCF Proxy",              "NOPAT = EBIT × (1−t)",              "NEE growth capex creates regulated assets; maintenance capex ≈ D&A"),
        ("Risk-Free Rate",         f"{RISK_FREE_RATE:.2%}",             "10-year US Treasury yield"),
        ("Equity Risk Premium",    f"{EQUITY_RISK_PREMIUM:.2%}",        "Damodaran US ERP"),
        ("Beta",                   f"{wacc_data['beta']:.2f}",          "5-year monthly, sourced from Yahoo Finance"),
        ("WACC",                   f"{wacc_data['wacc']:.2%}",          "Calculated from CAPM + after-tax cost of debt"),
        ("Terminal Growth Rate",   f"{TERMINAL_GROWTH_RATE:.2%}",       "Long-run regulated utility growth, in line with US GDP"),
        ("Revenue Growth (Yr1-2)", f"{REVENUE_GROWTH_RATES[0]:.0%}",    "Consistent with NEE 2024-2027 capital deployment plan"),
        ("Revenue Growth (Yr3-4)", f"{REVENUE_GROWTH_RATES[2]:.0%}",    "Step-down as rate base additions normalise"),
        ("Revenue Growth (Yr5)",   f"{REVENUE_GROWTH_RATES[4]:.0%}",    "Converging toward long-run growth"),
        ("EBIT Margin Target",     "34%",                               "Historical range 30-36%; assumes modest efficiency gains"),
        ("Effective Tax Rate",     f"{wacc_data['avg_tax']:.2%}",       "2-year avg; NEE benefits from ITC/PTC tax credits"),
        ("Shares Outstanding",     f"{hist['shares_out']/1e6:,.0f}M",   "Latest from Yahoo Finance"),
        ("Total Debt",             f"${hist['total_debt']/1e9:.1f}B",   "From most recent balance sheet"),
        ("Cash",                   f"${hist['cash']/1e9:.1f}B",         "Cash & equivalents"),
    ]

    for i, (param, val, note) in enumerate(notes):
        row_r = i + 2
        is_hdr = (i == 0)
        bg = DARK_GREEN if is_hdr else (LIGHT_GREEN if i % 2 == 0 else "FFFFFF")
        fg = "FFFFFF" if is_hdr else "000000"
        for col_i, val_i in enumerate([param, val, note], start=1):
            c = ws3.cell(row=row_r, column=col_i, value=val_i)
            c.font      = Font(bold=is_hdr, size=10, color=fg, name="Calibri")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border    = all_border()
        ws3.row_dimensions[row_r].height = 20

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 60

    wb.save(path)
    print(f"Saved → {path}")

# ── main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Fetching NEE financials…")
    hist = fetch_financials(TICKER)
    print(f"  Price: ${hist['current_price']:.2f}  |  Shares: {hist['shares_out']/1e6:.0f}M  |  Beta: {hist['beta']:.2f}")
    print(f"  Debt: ${hist['total_debt']/1e9:.1f}B  |  Cash: ${hist['cash']/1e9:.1f}B  |  Eff Tax: {hist['avg_tax']:.1%}")
    print(f"  Last EBIT: ${hist['ebit'][-1]/1e9:.1f}B  |  NOPAT: ${hist['nopat'][-1]/1e9:.1f}B")

    print("Building projections…")
    proj = build_projections(hist)

    print("Calculating WACC…")
    wacc_data = calc_wacc(hist)
    print(f"  WACC: {wacc_data['wacc']:.2%}  Ke: {wacc_data['ke']:.2%}  Kd: {wacc_data['kd']:.2%}")

    print("Running DCF…")
    dcf = calc_dcf(proj, wacc_data, hist)
    updown = (dcf["implied_price"] - hist["current_price"]) / hist["current_price"]
    print(f"  Implied: ${dcf['implied_price']:.2f}  Current: ${hist['current_price']:.2f}  ({updown:+.1%})")

    print("Building sensitivity table…")
    sens_table, sens_wacc, sens_tgr = sensitivity_table(proj, wacc_data["wacc"], hist)

    out = "/Users/carleesantel/Desktop/renewables-comps-analysis/nee-dcf/NEE_DCF_Model.xlsx"
    print("Writing Excel…")
    write_excel(hist, proj, wacc_data, dcf, sens_table, sens_wacc, sens_tgr, out)
    print("Done.")
